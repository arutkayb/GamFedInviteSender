from openpyxl import load_workbook
import util.invite_util as lang


class InvitePerson:
    name = ""
    mail = ""
    file_name = ""


class ExcelUtil:
    __workbook = ""
    __invite_sheet = ""

    def __init__(self, invite_file):
        self.__workbook = load_workbook(invite_file)
        self.__invite_sheet = self.__workbook.get_sheet_by_name('Davetiye')

    def get_invite_list(self):
        invite_list = []
        for i in range(2, 500):
            name = self.__invite_sheet.cell(row=i, column=1).value
            mail = self.__invite_sheet.cell(row=i, column=2).value
            send_status = self.__invite_sheet.cell(row=i, column=6).value

            if send_status is None and mail is not None:
                if name is None:
                    break
                else:
                    file_name = lang.get_eng_word(name)
                    invite = InvitePerson()
                    invite.name = name
                    invite.file_name = file_name
                    invite.mail = mail

                    invite_list.append(invite)

        return invite_list

    def set_column(self, column, value):
        pass
